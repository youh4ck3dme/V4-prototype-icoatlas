import json
from services.export_service import export_batch_to_excel
from openpyxl import load_workbook
import io

def test_export():
    sample_companies = [
        {
            "company_identifier": "54430178",
            "company_name": "BAMAT Service, s. r. o.",
            "country": "SK",
            "risk_score": 8.0,
            "company_data": {
                "ico": "54430178",
                "name": "BAMAT Service, s. r. o.",
                "street": "Čadečka 541",
                "city": "Čadca",
                "zip": "022 01",
                "dic": "1234567890",
                "ic_dph": "SK1234567890",
                "registration_id": "79245/L",
                "registration_section": "Sro",
                "capital": "5 000 EUR",
                "status": "Aktívna"
            }
        }
    ]
    
    excel_bytes = export_batch_to_excel(sample_companies)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print("Headers:", headers)
    
    row_data = [cell.value for cell in ws[2]]
    print("Row Data:", row_data)
    
    expected_headers = ["Street", "City", "Zip", "DIČ", "IČ DPH", "Vložka", "Oddiel", "Základné imanie"]
    # Check if they exist (translated or exact)
    # Our headers are in Slovak
    slovak_expected = ["Ulica", "Mesto", "PSČ", "DIČ", "IČ DPH", "Vložka", "Oddiel", "Základné imanie"]
    
    for h in slovak_expected:
        if h in headers:
            print(f"✅ Header '{h}' found.")
        else:
            print(f"❌ Header '{h}' MISSING.")

test_export()
